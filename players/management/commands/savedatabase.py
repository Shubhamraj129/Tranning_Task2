from django.core.management.base import BaseCommand
from players.models import *
from django.core.exceptions import ObjectDoesNotExist
import openpyxl
import ast


class Command(BaseCommand):

    def handle(self, *args, **options):
        book = openpyxl.load_workbook("/home/shubham/PycharmProjects/Scrapping_web/data11.xlsx")
        sheet = book.active
        fieldnames = ['Player_Name', 'Player_Image', 'Position', 'Height', 'Weight', 'HighSchool', 'City',
                      'Class', 'Offered_Clubs', 'Committed', 'Committed_College', 'Required_By',
                      'Twitter']  # sheet.max_row + 1
        for i in range(1, sheet.max_row):
            list2 = []
            for cell in sheet[i]:
                list2.append(cell.value)
            dictionary = dict(zip(fieldnames, list2))
            playername = dictionary.get('Player_Name')
            value = playername.split(' ')
            firstname = (value[0])
            lastname = (value[1])
            image = dictionary.get('Player_Image')
            height = dictionary.get('Height')
            weight = dictionary.get('Weight')
            if not Prospect.objects.filter(first_name=firstname, last_name=lastname, height=height, image=image):
                print(playername)
                position = dictionary.get('Position')
                try:
                    pos = Position.objects.get(pos=position)
                except ObjectDoesNotExist:
                    friend = Position(pos=position)
                    friend.save()
                position_id = Position.objects.get(pos=position)
                highSchool = dictionary.get('HighSchool')
                if highSchool == ' ':
                    print('')
                else:
                    try:
                        school = HighSchool.objects.get(name=highSchool)
                    except ObjectDoesNotExist:
                        friend = HighSchool(name=highSchool)
                        friend.save()
                highSchool_id = HighSchool.objects.get(name=highSchool)
                year = dictionary.get('Class')
                try:
                    years = Year.objects.get(year=year)
                except ObjectDoesNotExist:
                    friend = Year(year=year)
                    friend.save()
                class_id = Year.objects.get(year=dictionary.get('Class'))
                country1 = Country.objects.get_or_create()
                country_id = Country.objects.get(name='US')
                citystate = dictionary.get('City')
                print(citystate)
                csvalue = citystate.split(',')
                state = (csvalue[1])
                statename = State.objects.get_or_create(name=state, country=country_id)
                city = (csvalue[0])
                state_id = State.objects.get(name=state)
                cityname = City.objects.get_or_create(name=city, state=state_id)
                city_id = City.objects.get(name=city, state=state_id)
                offeredclubs = dictionary.get('Offered_Clubs')
                teams = ast.literal_eval(offeredclubs)
                offersid = Offers.objects.create()
                for logo, name in teams.items():
                    try:
                        friend = Team.objects.get(name=name)
                    except ObjectDoesNotExist:
                        friend = Team(name=name, logo=logo)
                        friend.save()
                for logo, name in teams.items():
                    teamid = Team.objects.get(name=name)
                    offersid.teams.add(teamid)
                offers_id = Offers.objects.get(pk=offersid.id)
                small = Prospect.objects.create(first_name=firstname, last_name=lastname, height=height, image=image,
                                                weight=weight, city=city_id, school=highSchool_id, position=position_id,
                                                offer=offers_id, year=class_id)
                player_id = Prospect.objects.get(first_name=firstname, last_name=lastname, height=height, image=image)
                if dictionary.get('Committed') == ' ':
                    print('')
                else:
                    committed = dictionary.get('Committed')
                    committedcollege = dictionary.get('Committed_College')
                    team_id = Team.objects.get(name=committedcollege)
                    required_By = dictionary.get('Required_By')
                    small = HardCommited.objects.create(commited=committed, requited_by=required_By, team=team_id,
                                                        player=player_id)
                twitterusername = dictionary.get('Twitter')
                if dictionary.get('Twitter') == ' ':
                    print("No Account")
                else:
                    TwitterInfo.objects.create(player_id=player_id, username=twitterusername)
